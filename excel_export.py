"""
excel_export.py  —  FundLens
Generates a polished, multi-sheet Excel workbook of the Top 50 mutual funds.
Suitable as a portfolio deliverable / for sharing with non-technical stakeholders.

Sheets produced:
  1. Top 50 Funds        — ranked table with conditional formatting
  2. Category Summary    — avg returns & expense ratios by category
  3. AMC Performance     — top AMCs by AUM with scoring
  4. Risk Cluster Report — cluster profile with labelled risk tiers
  5. Data Dictionary     — column descriptions for every metric used
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

INPUT_PATH = "data/processed/mutual_funds_scored.csv"
PBI_DIR    = "data/powerbi"
OUTPUT_PATH = "data/powerbi/FundLens_Top50_Report.xlsx"
TOP_N = 50

os.makedirs(PBI_DIR, exist_ok=True)

# ── Color palette (matching FundLens branding)
C_NAVY      = "0D3349"
C_TEAL      = "27C4A0"
C_ORANGE    = "F5A623"
C_RED       = "E84855"
C_LIGHTBLUE = "1A6B8A"
C_HEADER_BG = "0D3349"
C_HEADER_FG = "FFFFFF"
C_ALT_ROW   = "EFF6FA"
C_WHITE     = "FFFFFF"
C_BORDER    = "CBD5E1"
C_SCORE_HI  = "27C4A0"
C_SCORE_LO  = "E84855"

# ── Styles helpers
def hdr_font(bold=True, size=11, color=C_HEADER_FG):
    return Font(name="Arial", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="1E293B"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def hdr_fill(color=C_HEADER_BG):
    return PatternFill("solid", fgColor=color)

def alt_fill():
    return PatternFill("solid", fgColor=C_ALT_ROW)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, headers, row=1, bg=C_HEADER_BG, height=32):
    ws.row_dimensions[row].height = height
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = hdr_font()
        cell.fill = hdr_fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_rows(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        fill = alt_fill() if r % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = body_font()
            cell.border = thin_border()
            cell.alignment = left()

def set_col_widths(ws, widths: dict):
    """widths: {column_letter: width}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

# ══════════════════════════════════════════════════════════════
print("=" * 60)
print("  FUNDLENS — EXCEL EXPORT")
print("=" * 60)

df = pd.read_csv(INPUT_PATH)
print(f"  Loaded: {df.shape[0]:,} rows\n")

top_col = f"is_top{TOP_N}"
if top_col not in df.columns:
    thresh = df["investment_score"].nlargest(TOP_N).min()
    df[top_col] = df["investment_score"] >= thresh

top50 = (
    df[df[top_col] == True]
    .sort_values("investment_score", ascending=False)
    .reset_index(drop=True)
)
top50.insert(0, "rank", range(1, len(top50) + 1))

wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

# ══════════════════════════════════════════════════════════════
# SHEET 1 — TOP 50 FUNDS
# ══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Top 50 Funds")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"    # freeze header rows

# Title row
ws1.merge_cells("A1:N1")
title_cell = ws1["A1"]
title_cell.value = f"🏆  FundLens — Top {TOP_N} Mutual Fund Schemes by Investment Attractiveness Score"
title_cell.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
title_cell.fill = PatternFill("solid", fgColor=C_NAVY)
title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.row_dimensions[1].height = 36

T50_COLS = [c for c in [
    "rank", "scheme_name", "amc_name", "category",
    "fund_type", "risk_cluster_label",
    "returns_1yr", "returns_3yr", "returns_5yr",
    "expense_ratio", "sharpe", "alpha",
    "fund_size_cr", "investment_score"
] if c in top50.columns]

DISPLAY_HEADERS = {
    "rank":              "Rank",
    "scheme_name":       "Scheme Name",
    "amc_name":          "AMC",
    "category":          "Category",
    "fund_type":         "Fund Type",
    "risk_cluster_label":"Risk Tier",
    "returns_1yr":       "1Y Return (%)",
    "returns_3yr":       "3Y Return (%)",
    "returns_5yr":       "5Y Return (%)",
    "expense_ratio":     "Expense Ratio (%)",
    "sharpe":            "Sharpe Ratio",
    "alpha":             "Alpha",
    "fund_size_cr":      "Fund Size (Cr)",
    "investment_score":  "FundLens Score",
}

headers = [DISPLAY_HEADERS.get(c, c) for c in T50_COLS]
apply_header_row(ws1, headers, row=2, height=34)

for r_idx, (_, row) in enumerate(top50[T50_COLS].iterrows(), start=3):
    ws1.row_dimensions[r_idx].height = 22
    fill = alt_fill() if r_idx % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
    for c_idx, col in enumerate(T50_COLS, start=1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=row[col])
        cell.fill = fill
        cell.font = body_font(bold=(col == "scheme_name"))
        cell.border = thin_border()

        # Number formatting
        if col in ("returns_1yr", "returns_3yr", "returns_5yr", "expense_ratio", "sharpe", "alpha"):
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == "investment_score":
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", bold=True, size=10, color=C_NAVY)
        elif col == "fund_size_cr":
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col == "rank":
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
            cell.fill = PatternFill("solid", fgColor=C_LIGHTBLUE)
        else:
            cell.alignment = left()

# Conditional formatting: color scale on investment_score column
score_col_idx = T50_COLS.index("investment_score") + 1 if "investment_score" in T50_COLS else None
if score_col_idx:
    score_col_letter = get_column_letter(score_col_idx)
    score_range = f"{score_col_letter}3:{score_col_letter}{2 + len(top50)}"
    ws1.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="min", start_color=C_SCORE_LO,
            mid_type="percentile", mid_value=50, mid_color="F5A623",
            end_type="max", end_color=C_SCORE_HI
        )
    )

# Column widths
col_widths = {
    "A": 6, "B": 46, "C": 28, "D": 22, "E": 14, "F": 26,
    "G": 13, "H": 13, "I": 13, "J": 16, "K": 13, "L": 10,
    "M": 16, "N": 15,
}
set_col_widths(ws1, col_widths)

print(f"  ✔  Sheet 'Top {TOP_N} Funds'       → {len(top50)} rows")

# ══════════════════════════════════════════════════════════════
# SHEET 2 — CATEGORY SUMMARY
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Category Summary")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:H1")
ws2["A1"].value = "📊  Category-Level Performance Summary"
ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
ws2["A1"].fill = PatternFill("solid", fgColor=C_LIGHTBLUE)
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[1].height = 34

if "category" in df.columns:
    agg_d = {
        "Fund Count":         ("scheme_name", "count"),
        "Avg 1Y Return (%)":  ("returns_1yr", "mean"),
        "Avg 3Y Return (%)":  ("returns_3yr", "mean"),
        "Avg 5Y Return (%)":  ("returns_5yr", "mean"),
        "Avg Expense Ratio":  ("expense_ratio", "mean"),
        "Avg Sharpe":         ("sharpe", "mean"),
        "Avg FundLens Score": ("investment_score", "mean"),
    }
    cat_df = df.groupby("category").agg(**{
        k: v for k, v in agg_d.items()
    }).round(2).reset_index().sort_values("Avg 3Y Return (%)", ascending=False)

    headers2 = ["Category"] + list(agg_d.keys())
    apply_header_row(ws2, headers2, row=2, bg=C_LIGHTBLUE, height=30)

    for r_idx, (_, row) in enumerate(cat_df.iterrows(), start=3):
        ws2.row_dimensions[r_idx].height = 20
        fill = alt_fill() if r_idx % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
        ws2.cell(row=r_idx, column=1, value=row["category"]).fill = fill
        ws2.cell(row=r_idx, column=1).font = body_font(bold=True)
        ws2.cell(row=r_idx, column=1).border = thin_border()
        ws2.cell(row=r_idx, column=1).alignment = left()
        for c_idx, col in enumerate(agg_d.keys(), start=2):
            cell = ws2.cell(row=r_idx, column=c_idx, value=row[col])
            cell.fill = fill
            cell.font = body_font()
            cell.border = thin_border()
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Color scale on Avg 3Y Return
    ws2.conditional_formatting.add(
        f"C3:C{2 + len(cat_df)}",
        ColorScaleRule(start_type="min", start_color=C_SCORE_LO,
                       end_type="max", end_color=C_SCORE_HI)
    )

    set_col_widths(ws2, {"A": 30, "B": 12, "C": 16, "D": 16, "E": 16,
                         "F": 16, "G": 14, "H": 18})
    print(f"  ✔  Sheet 'Category Summary'     → {len(cat_df)} categories")

# ══════════════════════════════════════════════════════════════
# SHEET 3 — AMC PERFORMANCE
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("AMC Performance")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:G1")
ws3["A1"].value = "🏦  AMC Performance — Top 20 by AUM"
ws3["A1"].font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
ws3["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[1].height = 34

if "amc_name" in df.columns:
    top_col_local = f"is_top{TOP_N}"
    amc_agg = {
        "Fund Count":         ("scheme_name", "count"),
        "Total AUM (Cr)":     ("fund_size_cr", "sum"),
        "Avg Composite Ret":  ("composite_return", "mean"),
        "Avg Expense Ratio":  ("expense_ratio", "mean"),
        "Avg FundLens Score": ("investment_score", "mean"),
    }
    if top_col_local in df.columns:
        amc_agg[f"Top {TOP_N} Count"] = (top_col_local, "sum")

    amc_df = (
        df.groupby("amc_name").agg(**amc_agg)
        .round(2).reset_index()
        .nlargest(20, "Total AUM (Cr)").reset_index(drop=True)
    )
    amc_df.insert(0, "#", range(1, len(amc_df) + 1))

    headers3 = ["#", "AMC Name"] + list(amc_agg.keys())
    apply_header_row(ws3, headers3, row=2, bg=C_NAVY, height=30)

    for r_idx, (_, row) in enumerate(amc_df.iterrows(), start=3):
        ws3.row_dimensions[r_idx].height = 20
        fill = alt_fill() if r_idx % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
        ws3.cell(row=r_idx, column=1, value=row["#"]).fill = fill
        ws3.cell(row=r_idx, column=1).font = body_font(bold=True)
        ws3.cell(row=r_idx, column=1).border = thin_border()
        ws3.cell(row=r_idx, column=1).alignment = center()
        ws3.cell(row=r_idx, column=2, value=row["amc_name"]).fill = fill
        ws3.cell(row=r_idx, column=2).font = body_font(bold=True)
        ws3.cell(row=r_idx, column=2).border = thin_border()
        ws3.cell(row=r_idx, column=2).alignment = left()
        for c_idx, col in enumerate(amc_agg.keys(), start=3):
            cell = ws3.cell(row=r_idx, column=c_idx, value=row[col])
            cell.fill = fill
            cell.font = body_font()
            cell.border = thin_border()
            cell.number_format = '#,##0.00' if "AUM" in col else '0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")

    set_col_widths(ws3, {"A": 5, "B": 36, "C": 12, "D": 18, "E": 18, "F": 16, "G": 16, "H": 14})
    print(f"  ✔  Sheet 'AMC Performance'      → {len(amc_df)} AMCs")

# ══════════════════════════════════════════════════════════════
# SHEET 4 — RISK CLUSTER REPORT
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Risk Cluster Report")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A3"

ws4.merge_cells("A1:H1")
ws4["A1"].value = "⚖️  Risk Cluster Analysis — KMeans Segmentation"
ws4["A1"].font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
ws4["A1"].fill = PatternFill("solid", fgColor=C_TEAL.replace("27", "1A"))  # darker teal
ws4["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.row_dimensions[1].height = 34

if "risk_cluster_label" in df.columns:
    cluster_agg = (
        df.groupby("risk_cluster_label")
        .agg(
            fund_count        = ("scheme_name", "count"),
            avg_return        = ("composite_return", "mean"),
            avg_std_dev       = ("std_dev", "mean"),
            avg_sharpe        = ("sharpe", "mean"),
            avg_expense_ratio = ("expense_ratio", "mean"),
            avg_score         = ("investment_score", "mean"),
        )
        .round(3)
        .reset_index()
        .sort_values("avg_score", ascending=False)
    )

    headers4 = [
        "Risk Tier", "Fund Count", "Avg Composite Return (%)",
        "Avg Std Dev", "Avg Sharpe", "Avg Expense Ratio", "Avg FundLens Score"
    ]
    TIER_COLORS = ["27C4A0", "1A6B8A", "F5A623", "E84855", "0D3349"]
    apply_header_row(ws4, headers4, row=2, bg="1A5F7A", height=30)

    for r_idx, (_, row) in enumerate(cluster_agg.iterrows(), start=3):
        ws4.row_dimensions[r_idx].height = 26
        tier_color = TIER_COLORS[(r_idx - 3) % len(TIER_COLORS)]
        fill = PatternFill("solid", fgColor="F0FAFA") if r_idx % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
        for c_idx, col in enumerate(cluster_agg.columns, start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=row[col])
            cell.fill = fill
            cell.border = thin_border()
            if c_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=10, color=tier_color)
                cell.alignment = left()
            else:
                cell.font = body_font()
                cell.number_format = '0.00'
                cell.alignment = center()

    set_col_widths(ws4, {"A": 32, "B": 13, "C": 24, "D": 13, "E": 13, "F": 18, "G": 20})
    print(f"  ✔  Sheet 'Risk Cluster Report'  → {len(cluster_agg)} clusters")

# ══════════════════════════════════════════════════════════════
# SHEET 5 — DATA DICTIONARY
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Data Dictionary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:D1")
ws5["A1"].value = "📖  FundLens Data Dictionary — Column Definitions"
ws5["A1"].font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
ws5["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
ws5["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws5.row_dimensions[1].height = 34

DICTIONARY = [
    ("scheme_name",       "Text",    "Full name of the mutual fund scheme"),
    ("amc_name",          "Text",    "Asset Management Company managing the fund"),
    ("category",          "Text",    "SEBI-defined broad category (e.g., Equity, Debt, Hybrid)"),
    ("sub_category",      "Text",    "Narrower classification within category"),
    ("fund_type",         "Text",    "Open-ended / Close-ended / Interval"),
    ("risk_level",        "Text",    "Self-declared risk level by AMC"),
    ("rating",            "Numeric", "Fund star rating (1–5)"),
    ("returns_1yr",       "Numeric", "Trailing 1-year return (%)"),
    ("returns_3yr",       "Numeric", "Trailing 3-year CAGR (%)"),
    ("returns_5yr",       "Numeric", "Trailing 5-year CAGR (%)"),
    ("composite_return",  "Numeric", "Weighted blend: 20% × 1yr + 35% × 3yr + 45% × 5yr"),
    ("expense_ratio",     "Numeric", "Annual expense ratio charged to investors (%)"),
    ("net_return",        "Numeric", "composite_return minus expense_ratio"),
    ("sharpe",            "Numeric", "Sharpe ratio — excess return per unit of total risk"),
    ("alpha",             "Numeric", "Excess return over benchmark (Jensen's Alpha)"),
    ("beta",              "Numeric", "Sensitivity to benchmark market movements"),
    ("std_dev",           "Numeric", "Annualised standard deviation of returns (volatility proxy)"),
    ("rari",              "Numeric", "Risk-Adjusted Return Index = composite_return / std_dev"),
    ("fund_size_cr",      "Numeric", "Assets Under Management in Indian Crore (₹)"),
    ("fund_size_category","Text",    "AUM bucket: Small <500Cr, Mid 500–5KCr, Large 5K–20KCr, Giant >20KCr"),
    ("min_sip",           "Numeric", "Minimum SIP instalment amount (₹)"),
    ("min_lumpsum",       "Numeric", "Minimum one-time investment amount (₹)"),
    ("return_consistency","Numeric", "Higher = more consistent returns across 1/3/5yr horizons"),
    ("cluster",           "Integer", "KMeans cluster assignment (0-indexed)"),
    ("risk_cluster_label","Text",    "Human-readable risk tier derived from cluster profiling"),
    ("investment_score",  "Numeric", "FundLens proprietary score 0–100. Weights: 3Y Return 30%, Composite 20%, Low Expense 20%, Sharpe 15%, Alpha 10%, RARI 5%"),
    ("rank_by_score",     "Integer", "Rank across all ~2,500 funds by investment_score (1 = best)"),
    ("rank_by_return",    "Integer", "Rank by composite_return"),
    ("rank_by_sharpe",    "Integer", "Rank by sharpe ratio"),
    (f"is_top{TOP_N}",    "Boolean", f"TRUE if fund is in the Top {TOP_N} by investment_score"),
]

dict_headers = ["Column Name", "Data Type", "Description"]
apply_header_row(ws5, dict_headers, row=2, bg=C_NAVY, height=30)

for r_idx, (col, dtype, desc) in enumerate(DICTIONARY, start=3):
    ws5.row_dimensions[r_idx].height = 20
    fill = alt_fill() if r_idx % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
    for c_idx, val in enumerate([col, dtype, desc], start=1):
        cell = ws5.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = fill
        cell.border = thin_border()
        cell.font = body_font(bold=(c_idx == 1), color=C_NAVY if c_idx == 1 else "1E293B")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws5.row_dimensions[r_idx].height = 22

set_col_widths(ws5, {"A": 25, "B": 12, "C": 80})
print(f"  ✔  Sheet 'Data Dictionary'      → {len(DICTIONARY)} entries")

# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"\n  💾  Saved → {OUTPUT_PATH}")
print("\n" + "=" * 60)
print("  ✅  Excel workbook complete — FundLens_Top50_Report.xlsx")
print("=" * 60)